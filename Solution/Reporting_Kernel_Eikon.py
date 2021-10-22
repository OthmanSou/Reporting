# -*- coding: utf-8 -*-
"""
Created on Wed Jul 21 15:40:15 2021

@author: osoua
"""

# -*- coding: utf-8 -*-
"""
Reporting v 0.0
"""

import eikon as ek
import pandas as pd
import refinitiv.dataplatform as rdp 
import docx
import numpy as np
import matplotlib.pyplot as plt
import QuantLib as ql
from datetime import datetime as dt
import openpyxl as xl


def ql_to_string(d):
	month = str(d.month())
	if d.month()<10: 
		month ='0' + month
	day = str(d.dayOfMonth())
	if d.dayOfMonth()<10:
		day ='0'+ day
	return str(d.year()) + '-' + month + '-' + day

def get_previous_friday_adj(d):
	Maturity_ = ql.Period(-1, ql.Days)
	Calendar_ = ql.TARGET()
	Convention_ = ql.ModifiedFollowing
	weekday = d.weekday()
	qlPreious_date_ = d
	while weekday!= 6:	
		qlPreious_date_ = Calendar_.advance(qlPreious_date_, Maturity_, Convention_, False)
		weekday = qlPreious_date_.weekday()
	r =ql_to_string(qlPreious_date_)
	return r
def delete_worksheet(file_workbook, SheetName):
	wb = xl.load_workbook(file_workbook)
	if SheetName in wb.sheetnames:
		wb.remove(wb[SheetName])
	wb.save(file_workbook)
	
def report_to_excel(dictio, ud, FileName, SheetName, ParamDataFrame, ParamSheetName):
	
	ParamDataFrame = param_data_frame
	book = xl.load_workbook(FileName)
	if SheetName in book.sheetnames:
			idx = book.sheetnames.index(SheetName)
			book.remove(book[SheetName])
			book.create_sheet(SheetName, idx)
			idx = book.sheetnames.index(ParamSheetName)
			book.remove(book[ParamSheetName])
			book.create_sheet(ParamSheetName, idx)
	book.save(FileName)

	book = xl.load_workbook(FileName)
	with pd.ExcelWriter(FileName, engine='openpyxl') as writer:
		
		writer.book = book
		writer.sheets = dict((ws.title, ws) for ws in book.worksheets)    
		
	    ## Your dataframe to append.     
		row_counter = 0
		df = dictio[ud[0]]
		for i in range(len(ud)):
			df = dictio[ud[i]]
			df.to_excel(writer, sheet_name = SheetName, startrow = row_counter, startcol =0)
			row_counter = row_counter + len(dictio[ud[i]])+1
		ParamDataFrame.to_excel(writer, sheet_name = ParamSheetName)
		writer.save()	
		
	

	   
def report_to_word(df):
	df = df.reset_index(inplace = False)
	# open an existing document
	doc = docx.Document()
	# add a table to the end and create a reference variable
	# extra row is so we can add the header row
	t = doc.add_table(df.shape[0]+1, df.shape[1])

	# add the header rows.
	for j in range(df.shape[-1]):
		t.cell(0,j).text = df.columns[j]

	# add the rest of the data frame
	for i in range(df.shape[0]):
	    for j in range(df.shape[-1]):
	        t.cell(i+1,j).text = str(df.values[i,j])

	# save the doc
	doc.save(WordName)
   

#variables and arguments
App_key = r'310a8e0605ee4dbe9befd54b7911c060428878ae'
WordName = r'C:\Users\osoua\Desktop\Work\Reporting Financier\Solution\Reporting.docx'
File_name = r'C:\Users\osoua\Desktop\Work\Reporting Financier\Solution\Reporting.xlsm'
ResultsSheetName = 'Results'
ParamSheetName ='Param'
TemplateName = r'C:\Users\osoua\Desktop\Work\Reporting Financier\Solution\ReportingTemplates.xlsx'
ek.set_app_key(App_key)
rdp.open_desktop_session(App_key)



param_data_frame = pd.read_excel(File_name,sheet_name = 'Main', skiprows=5, usecols='G:I')
Report_number = len(param_data_frame)

Underlyings = []
time_series = {}
reportingDF = {}
n = 20

for Pos in range(Report_number):
	Start_date = "2019-01-01"
	End_date = param_data_frame.iloc[Pos]['Analysis Date'].strftime('%Y-%m-%d')
	Underlying = [param_data_frame.iloc[Pos]['Underlyings']]
	Interval = param_data_frame.iloc[Pos]['Time Frame']
	
	#time_series[Underlying[0]] = pd.read_excel(r'C:\Users\osoua\Desktop\Work\Reporting Financier\Solution\TestCAC.xlsx', index_col =0)
	time_series[Underlying[0]] = ek.get_timeseries(Underlying, 
									start_date = Start_date,
								    end_date = End_date, 
									interval = Interval)
	#Typical Price, TP = (High + Low + Close) /3
	time_series[Underlying[0]]['TP'] = (time_series[Underlying[0]]['HIGH'] +
										time_series[Underlying[0]]['LOW']
										+ time_series[Underlying[0]]['CLOSE'])/3
	#VOL Bollinger
	time_series[Underlying[0]]['VOLB'] = time_series[Underlying[0]]['TP'].rolling(n).std()
	
	#Previous Close
	time_series[Underlying[0]]['P_Close'] = time_series[Underlying[0]]['CLOSE'].shift(periods=1)
	
	#True Range, TR = max(High-Low, High-PreviousClose, PreviousClose-Low)
	time_series[Underlying[0]]['H-L'] = time_series[Underlying[0]]['HIGH']-time_series[Underlying[0]]['LOW']
	time_series[Underlying[0]]['H-PC'] = time_series[Underlying[0]]['HIGH']-time_series[Underlying[0]]['P_Close']
	time_series[Underlying[0]]['PC-L'] = time_series[Underlying[0]]['P_Close']-time_series[Underlying[0]]['LOW']
	time_series[Underlying[0]]['TR'] = time_series[Underlying[0]][['H-L','H-PC','PC-L']].max(axis = 1)
	time_series[Underlying[0]].drop(['P_Close','H-L','H-PC','PC-L'], axis =1, inplace=True)
	
	#Average True Range, or VOLK ATR = TR's exponential moving average
	time_series[Underlying[0]]['VOLK'] = time_series[Underlying[0]]['TR'].ewm(span=n, adjust=False).mean()
	time_series[Underlying[0]].drop(['TR'], axis =1, inplace=True)
	
	#Import the reporting DataFrame format and fill It
	reportingDF[Underlying[0]] = pd.read_excel(File_name, sheet_name = 'Reporting', skiprows=5, usecols='F:F', index_col =0)
	reportingDF[Underlying[0]].index.name =  Underlying[0]
	
	#Volume
	reportingDF[Underlying[0]].at['Volume','Spot'] = time_series[Underlying[0]].at[End_date,'VOLUME']
	
	#Spot 
	reportingDF[Underlying[0]].at['Open','Spot'] = time_series[Underlying[0]].at[End_date,'OPEN']
	reportingDF[Underlying[0]].at['Close','Spot'] = time_series[Underlying[0]].at[End_date,'CLOSE']
	reportingDF[Underlying[0]].at['High','Spot'] = time_series[Underlying[0]].at[End_date,'HIGH']
	reportingDF[Underlying[0]].at['Low','Spot'] = time_series[Underlying[0]].at[End_date,'LOW']
	reportingDF[Underlying[0]].at['VolK','Spot'] = time_series[Underlying[0]].at[End_date,'VOLK']
	reportingDF[Underlying[0]].at['VolB','Spot'] = time_series[Underlying[0]].at[End_date,'VOLB']
	
	#Previous Week % Change
	Calendar = ql.TARGET()
	Maturity = ql.Period(-1, ql.Weeks)
	Convention = ql.ModifiedFollowing
	QlEnd_date = ql.Date(End_date, '%Y-%m-%d')
	QlPreious_week_date = Calendar.advance(QlEnd_date, Maturity, Convention, False)
	Preious_week_date = ql_to_string(QlPreious_week_date)
	
	reportingDF[Underlying[0]].at['Open','Previous Week % Change']  = 100*(time_series[Underlying[0]].at[End_date,'OPEN'] - time_series[Underlying[0]].at[Preious_week_date,'OPEN'])/time_series[Underlying[0]].at[Preious_week_date,'OPEN']
	reportingDF[Underlying[0]].at['Close','Previous Week % Change'] = 100*(time_series[Underlying[0]].at[End_date,'CLOSE'] - time_series[Underlying[0]].at[Preious_week_date,'CLOSE'])/time_series[Underlying[0]].at[Preious_week_date,'CLOSE']
	reportingDF[Underlying[0]].at['High','Previous Week % Change']  = 100*(time_series[Underlying[0]].at[End_date,'HIGH'] - time_series[Underlying[0]].at[Preious_week_date,'HIGH'])/time_series[Underlying[0]].at[Preious_week_date,'HIGH']
	reportingDF[Underlying[0]].at['Low','Previous Week % Change']   = 100*(time_series[Underlying[0]].at[End_date,'LOW'] - time_series[Underlying[0]].at[Preious_week_date,'LOW'])/time_series[Underlying[0]].at[Preious_week_date,'LOW']
	reportingDF[Underlying[0]].at['VolK','Previous Week % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLK'] - time_series[Underlying[0]].at[Preious_week_date,'VOLK'])/time_series[Underlying[0]].at[Preious_week_date,'VOLK']
	reportingDF[Underlying[0]].at['VolB','Previous Week % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLB'] - time_series[Underlying[0]].at[Preious_week_date,'VOLB'])/time_series[Underlying[0]].at[Preious_week_date,'VOLB']
	
	#Previous Year % Change
	Calendar = ql.TARGET()
	Maturity = ql.Period(-1, ql.Years)
	Convention = ql.ModifiedFollowing
	QlEnd_date = ql.Date(End_date, '%Y-%m-%d')
	QlPreious_week_date = Calendar.advance(QlEnd_date, Maturity, Convention, False)
	Preious_week_date = ql_to_string(QlPreious_week_date)
	
	reportingDF[Underlying[0]].at['Open','Previous Year % Change']  = 100*(time_series[Underlying[0]].at[End_date,'OPEN'] - time_series[Underlying[0]].at[Preious_week_date,'OPEN'])/time_series[Underlying[0]].at[Preious_week_date,'OPEN']
	reportingDF[Underlying[0]].at['Close','Previous Year % Change'] = 100*(time_series[Underlying[0]].at[End_date,'CLOSE'] - time_series[Underlying[0]].at[Preious_week_date,'CLOSE'])/time_series[Underlying[0]].at[Preious_week_date,'CLOSE']
	reportingDF[Underlying[0]].at['High','Previous Year % Change']  = 100*(time_series[Underlying[0]].at[End_date,'HIGH'] - time_series[Underlying[0]].at[Preious_week_date,'HIGH'])/time_series[Underlying[0]].at[Preious_week_date,'HIGH']
	reportingDF[Underlying[0]].at['Low','Previous Year % Change']   = 100*(time_series[Underlying[0]].at[End_date,'LOW'] - time_series[Underlying[0]].at[Preious_week_date,'LOW'])/time_series[Underlying[0]].at[Preious_week_date,'LOW']
	reportingDF[Underlying[0]].at['VolK','Previous Year % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLK'] - time_series[Underlying[0]].at[Preious_week_date,'VOLK'])/time_series[Underlying[0]].at[Preious_week_date,'VOLK']
	reportingDF[Underlying[0]].at['VolB','Previous Year % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLB'] - time_series[Underlying[0]].at[Preious_week_date,'VOLB'])/time_series[Underlying[0]].at[Preious_week_date,'VOLB']
	
	
	reportingDF[Underlying[0]].at['Open','Previous 5 Year % Change']  = 100*(time_series[Underlying[0]].at[End_date,'OPEN'] - time_series[Underlying[0]].at[Preious_week_date,'OPEN'])/time_series[Underlying[0]].at[Preious_week_date,'OPEN']
	reportingDF[Underlying[0]].at['Close','Previous 5 Year % Change'] = 100*(time_series[Underlying[0]].at[End_date,'CLOSE'] - time_series[Underlying[0]].at[Preious_week_date,'CLOSE'])/time_series[Underlying[0]].at[Preious_week_date,'CLOSE']
	reportingDF[Underlying[0]].at['High','Previous 5 Year % Change']  = 100*(time_series[Underlying[0]].at[End_date,'HIGH'] - time_series[Underlying[0]].at[Preious_week_date,'HIGH'])/time_series[Underlying[0]].at[Preious_week_date,'HIGH']
	reportingDF[Underlying[0]].at['Low','Previous 5 Year % Change']   = 100*(time_series[Underlying[0]].at[End_date,'LOW'] - time_series[Underlying[0]].at[Preious_week_date,'LOW'])/time_series[Underlying[0]].at[Preious_week_date,'LOW']
	reportingDF[Underlying[0]].at['VolK','Previous 5 Year % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLK'] - time_series[Underlying[0]].at[Preious_week_date,'VOLK'])/time_series[Underlying[0]].at[Preious_week_date,'VOLK']
	reportingDF[Underlying[0]].at['VolB','Previous 5 Year % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLB'] - time_series[Underlying[0]].at[Preious_week_date,'VOLB'])/time_series[Underlying[0]].at[Preious_week_date,'VOLB']
	
	
	#MTD % Change
	Maturity = ql.Period(-1, ql.Days)
	QlFirstOfMonth_date_adjusted = Calendar.advance(ql.Date(1,QlEnd_date.month(),QlEnd_date.year()), Maturity, Convention, False)
	FirstOfMonth_date_adjusted = ql_to_string(QlFirstOfMonth_date_adjusted)
	
	reportingDF[Underlying[0]].at['Open','MTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'OPEN'] - time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'OPEN'])/time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'OPEN']
	reportingDF[Underlying[0]].at['Close','MTD % Change'] = 100*(time_series[Underlying[0]].at[End_date,'CLOSE'] - time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'CLOSE'])/time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'CLOSE']
	reportingDF[Underlying[0]].at['High','MTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'HIGH'] - time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'HIGH'])/time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'HIGH']
	reportingDF[Underlying[0]].at['Low','MTD % Change']   = 100*(time_series[Underlying[0]].at[End_date,'LOW'] - time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'LOW'])/time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'LOW']
	reportingDF[Underlying[0]].at['VolK','MTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLK'] - time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'VOLK'])/time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'VOLK']
	reportingDF[Underlying[0]].at['VolB','MTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLB'] - time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'VOLB'])/time_series[Underlying[0]].at[FirstOfMonth_date_adjusted,'VOLB']
	
	#WTD % Change
	PreviousFriday_adjusted = get_previous_friday_adj(QlEnd_date)
	
	reportingDF[Underlying[0]].at['Open','WTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'OPEN'] - time_series[Underlying[0]].at[PreviousFriday_adjusted,'OPEN'])/time_series[Underlying[0]].at[PreviousFriday_adjusted,'OPEN']
	reportingDF[Underlying[0]].at['Close','WTD % Change'] = 100*(time_series[Underlying[0]].at[End_date,'CLOSE'] - time_series[Underlying[0]].at[PreviousFriday_adjusted,'CLOSE'])/time_series[Underlying[0]].at[PreviousFriday_adjusted,'CLOSE']
	reportingDF[Underlying[0]].at['High','WTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'HIGH'] - time_series[Underlying[0]].at[PreviousFriday_adjusted,'HIGH'])/time_series[Underlying[0]].at[PreviousFriday_adjusted,'HIGH']
	reportingDF[Underlying[0]].at['Low','WTD % Change']   = 100*(time_series[Underlying[0]].at[End_date,'LOW'] - time_series[Underlying[0]].at[PreviousFriday_adjusted,'LOW'])/time_series[Underlying[0]].at[PreviousFriday_adjusted,'LOW']
	reportingDF[Underlying[0]].at['VolK','WTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLK'] - time_series[Underlying[0]].at[PreviousFriday_adjusted,'VOLK'])/time_series[Underlying[0]].at[PreviousFriday_adjusted,'VOLK']
	reportingDF[Underlying[0]].at['VolB','WTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLB'] - time_series[Underlying[0]].at[PreviousFriday_adjusted,'VOLB'])/time_series[Underlying[0]].at[PreviousFriday_adjusted,'VOLB']
	
	#YTD % Change
	QlNewYear_date_adjusted = Calendar.advance(ql.Date(1,1,QlEnd_date.year()), Maturity, Convention, False)
	NewYear_date_adjusted = ql_to_string(QlNewYear_date_adjusted)
	
	reportingDF[Underlying[0]].at['Open','YTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'OPEN'] - time_series[Underlying[0]].at[NewYear_date_adjusted,'OPEN'])/time_series[Underlying[0]].at[NewYear_date_adjusted,'OPEN']
	reportingDF[Underlying[0]].at['Close','YTD % Change'] = 100*(time_series[Underlying[0]].at[End_date,'CLOSE'] - time_series[Underlying[0]].at[NewYear_date_adjusted,'CLOSE'])/time_series[Underlying[0]].at[NewYear_date_adjusted,'CLOSE']
	reportingDF[Underlying[0]].at['High','YTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'HIGH'] - time_series[Underlying[0]].at[NewYear_date_adjusted,'HIGH'])/time_series[Underlying[0]].at[NewYear_date_adjusted,'HIGH']
	reportingDF[Underlying[0]].at['Low','YTD % Change']   = 100*(time_series[Underlying[0]].at[End_date,'LOW'] - time_series[Underlying[0]].at[NewYear_date_adjusted,'LOW'])/time_series[Underlying[0]].at[NewYear_date_adjusted,'LOW']
	reportingDF[Underlying[0]].at['VolK','YTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLK'] - time_series[Underlying[0]].at[NewYear_date_adjusted,'VOLK'])/time_series[Underlying[0]].at[NewYear_date_adjusted,'VOLK']
	reportingDF[Underlying[0]].at['VolB','YTD % Change']  = 100*(time_series[Underlying[0]].at[End_date,'VOLB'] - time_series[Underlying[0]].at[NewYear_date_adjusted,'VOLB'])/time_series[Underlying[0]].at[NewYear_date_adjusted,'VOLB']
	Underlyings.append(Underlying[0])
	
	#Save in Word
	#report_to_word(reportingDF[Underlying[0]])
	
#save in excel
report_to_excel(reportingDF, Underlyings, TemplateName, ResultsSheetName, param_data_frame, ParamSheetName)
#Display




	
	
	
 
	
